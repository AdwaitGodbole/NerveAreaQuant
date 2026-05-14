import pandas as pd
import numpy as np
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilenames
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Hide main Tkinter window
Tk().withdraw()

# Step 1: Let user select TSV files
tsv_files = askopenfilenames(title="Select TSV files", filetypes=[("TSV files", "*.tsv")])

# Quadrant assignment function
def assign_quadrant(angle):
    if pd.isna(angle):
        return pd.NA
    elif 45 <= angle < 135:
        return "Quadrant 1"
    elif 135 <= angle < 225:
        return "Quadrant 4"
    elif 225 <= angle < 315:
        return "Quadrant 3"
    else:
        return "Quadrant 2"  # 315–360 and 0–45

# Labels for bins and top rows
quadrant_labels = ["Quadrant 1", "Quadrant 4", "Quadrant 3", "Quadrant 2", "Total"]

for tsv_file in tsv_files:
    folder = os.path.dirname(tsv_file)
    base_name = os.path.splitext(os.path.basename(tsv_file))[0]

    df = pd.read_csv(tsv_file, sep="\t")

    target_folder = os.path.join(folder, "TSV and Excel")
    os.makedirs(target_folder, exist_ok=True)

    shutil.move(tsv_file, os.path.join(target_folder, os.path.basename(tsv_file)))
    excel_path = os.path.join(target_folder, f"{base_name}.xlsx")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        # Raw Data sheet
        df.to_excel(writer, sheet_name="Raw Data", index=False)

        # Column copy map
        column_map = {
            "Name": "Name",
            "Centre X (Lumen)": "Centre X (Lumen) (μm)",
            "Centre Y (Lumen)": "Centre Y (Lumen) (μm)",
            "Centre X (TH negative area)": "Centre X (TH negative area) (μm)",
            "Centre X (TH positive area)": "Centre X (TH positive area) (μm)",
            "Centre Y (TH negative area)": "Centre Y (TH negative area) (μm)",
            "Centre Y (TH positive area)": "Centre Y (TH positive area) (μm)",
            "Area (TH negative area)": "Area (TH negative area) (μm²)",
            "Area (TH positive area)": "Area (TH positive area) (μm²)"
        }

        df_processed = pd.DataFrame()
        for src, dst in column_map.items():
            df_processed[dst] = df[src] if src in df.columns else pd.NA

        # Insert empty columns
        new_columns = [
            "Shifted X TH positive area (μm)",
            "Shifted Y TH positive area (μm)",
            "Shifted X TH negative area (μm)",
            "Shifted Y TH negative area (μm)",
            "Angle between TH positive area and lumen (rad)",
            "Angle between TH negative area and lumen (rad)",
            "Angle between TH positive area and lumen (deg)",
            "Angle between TH negative area and lumen (deg)",
            "Quadrants_TH positive area",
            "Quadrants_TH negative area",
            "Quadrant Bins_TH positive area",
            "Area per Quadrant Bin (μm²)_TH positive area",
            "Quadrant Bins_TH negative area",
            "Area per Quadrant Bin (μm²)_TH negative area",
            "Rounded off Area_TH positive (mm²) per Quadrant",
            "Rounded off Area_TH negative (mm²) per Quadrant",
            "Total Nerve Tissue (mm²) per Quadrant",
            "%TH positive nerve tissue per Quadrant",
            "%TH negative nerve tissue per Quadrant"
        ]

        insert_idx = df_processed.columns.get_loc("Area (TH positive area) (μm²)")
        for i, col in enumerate(new_columns):
            df_processed.insert(insert_idx + 1 + i, col, "")

        x_lumen = "Centre X (Lumen) (μm)"
        y_lumen = "Centre Y (Lumen) (μm)"
        x_th_pos = "Centre X (TH positive area) (μm)"
        y_th_pos = "Centre Y (TH positive area) (μm)"
        x_th_neg = "Centre X (TH negative area) (μm)"
        y_th_neg = "Centre Y (TH negative area) (μm)"

        lumen_x = pd.to_numeric(df_processed.loc[0, x_lumen], errors="coerce")
        lumen_y = pd.to_numeric(df_processed.loc[0, y_lumen], errors="coerce")

        df_processed["Shifted X TH positive area (μm)"] = pd.to_numeric(df_processed[x_th_pos], errors="coerce") - lumen_x
        df_processed["Shifted Y TH positive area (μm)"] = pd.to_numeric(df_processed[y_th_pos], errors="coerce") - lumen_y
        df_processed["Shifted X TH negative area (μm)"] = pd.to_numeric(df_processed[x_th_neg], errors="coerce") - lumen_x
        df_processed["Shifted Y TH negative area (μm)"] = pd.to_numeric(df_processed[y_th_neg], errors="coerce") - lumen_y

        df_processed["Angle between TH positive area and lumen (rad)"] = np.arctan2(
            df_processed["Shifted Y TH positive area (μm)"],
            df_processed["Shifted X TH positive area (μm)"]
        )
        df_processed["Angle between TH positive area and lumen (deg)"] = (
            np.degrees(df_processed["Angle between TH positive area and lumen (rad)"]) % 360
        )

        df_processed["Angle between TH negative area and lumen (rad)"] = np.arctan2(
            df_processed["Shifted Y TH negative area (μm)"],
            df_processed["Shifted X TH negative area (μm)"]
        )
        df_processed["Angle between TH negative area and lumen (deg)"] = (
            np.degrees(df_processed["Angle between TH negative area and lumen (rad)"]) % 360
        )

        df_processed["Quadrants_TH positive area"] = df_processed[
            "Angle between TH positive area and lumen (deg)"
        ].apply(assign_quadrant)

        df_processed["Quadrants_TH negative area"] = df_processed[
            "Angle between TH negative area and lumen (deg)"
        ].apply(assign_quadrant)

        for i, label in enumerate(quadrant_labels):
            if i < len(df_processed):
                df_processed.at[i, "Quadrant Bins_TH positive area"] = label
                df_processed.at[i, "Quadrant Bins_TH negative area"] = label

        area_pos = df_processed.groupby("Quadrants_TH positive area")[
            "Area (TH positive area) (μm²)"
        ].sum()
        for i, q in enumerate(quadrant_labels[:-1]):
            df_processed.at[i, "Area per Quadrant Bin (μm²)_TH positive area"] = area_pos.get(q, 0)
        if len(df_processed) > 4:
            df_processed.at[4, "Area per Quadrant Bin (μm²)_TH positive area"] = area_pos.sum()

        area_neg = df_processed.groupby("Quadrants_TH negative area")[
            "Area (TH negative area) (μm²)"
        ].sum()
        for i, q in enumerate(quadrant_labels[:-1]):
            df_processed.at[i, "Area per Quadrant Bin (μm²)_TH negative area"] = area_neg.get(q, 0)
        if len(df_processed) > 4:
            df_processed.at[4, "Area per Quadrant Bin (μm²)_TH negative area"] = area_neg.sum()

        df_processed.to_excel(writer, sheet_name="Processed Data", index=False)

    wb = load_workbook(excel_path)
    ws = wb["Processed Data"]

    round_pos_idx = df_processed.columns.get_loc("Rounded off Area_TH positive (mm²) per Quadrant")
    round_neg_idx = df_processed.columns.get_loc("Rounded off Area_TH negative (mm²) per Quadrant")
    area_pos_idx = df_processed.columns.get_loc("Area per Quadrant Bin (μm²)_TH positive area")
    area_neg_idx = df_processed.columns.get_loc("Area per Quadrant Bin (μm²)_TH negative area")
    total_nerve_idx = df_processed.columns.get_loc("Total Nerve Tissue (mm²) per Quadrant")
    pct_pos_idx = df_processed.columns.get_loc("%TH positive nerve tissue per Quadrant")
    pct_neg_idx = df_processed.columns.get_loc("%TH negative nerve tissue per Quadrant")

    for i in range(5):
        row_num = i + 2

        area_pos_letter = get_column_letter(area_pos_idx + 1)
        area_neg_letter = get_column_letter(area_neg_idx + 1)
        round_pos_letter = get_column_letter(round_pos_idx + 1)
        round_neg_letter = get_column_letter(round_neg_idx + 1)
        total_letter = get_column_letter(total_nerve_idx + 1)
        pct_pos_letter = get_column_letter(pct_pos_idx + 1)
        pct_neg_letter = get_column_letter(pct_neg_idx + 1)

        ws[f"{round_pos_letter}{row_num}"] = f"=ROUND({area_pos_letter}{row_num}/1000000,2)"
        ws[f"{round_neg_letter}{row_num}"] = f"=ROUND({area_neg_letter}{row_num}/1000000,2)"
        ws[f"{total_letter}{row_num}"] = f"={round_pos_letter}{row_num}+{round_neg_letter}{row_num}"

        ws[f"{pct_pos_letter}{row_num}"] = (
            f'=IF({total_letter}{row_num}=0,"NA",ROUND({round_pos_letter}{row_num}/{total_letter}{row_num}*100,2))'
        )
        ws[f"{pct_neg_letter}{row_num}"] = (
            f'=IF({total_letter}{row_num}=0,"NA",ROUND({round_neg_letter}{row_num}/{total_letter}{row_num}*100,2))'
        )

    for sheet in ["Raw Data", "Processed Data"]:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.font = Font(bold=False)
            cell.border = Border()
            cell.fill = PatternFill()
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 2

    ws = wb["Processed Data"]
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center")

    wb.save(excel_path)

print("Done: full pipeline complete with corrected mm² headers.")
